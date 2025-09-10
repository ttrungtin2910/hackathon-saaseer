"""
Chương trình đánh giá hiệu suất trích xuất thông tin hóa đơn/hợp đồng của AI model
Author: Assistant
Date: 2024
"""

import os
import time
import json
import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Tuple
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import io
from PIL import Image
from pdf2image import convert_from_bytes

# Import các module từ dự án hiện tại
from app.services.langgraph import langgraph_service
from app.models.contract import ExtractedContractData

# Cấu hình logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('evaluation_log.txt', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class ContractExtractionEvaluator:
    """Class đánh giá hiệu suất trích xuất thông tin hợp đồng"""
    
    def __init__(self, data_folder: str = "data"):
        self.data_folder = Path(data_folder)
        self.results = []
        self.performance_metrics = {
            'total_files': 0,
            'successful_extractions': 0,
            'failed_extractions': 0,
            'total_processing_time': 0,
            'average_processing_time': 0,
            'success_rate': 0
        }
        
        # Kiểm tra xem thư mục data có tồn tại không
        if not self.data_folder.exists():
            raise FileNotFoundError(f"Thư mục {data_folder} không tồn tại!")
            
        logger.info(f"Khởi tạo evaluator cho thư mục: {self.data_folder}")
    
    def get_file_list(self) -> List[Path]:
        """Lấy danh sách tất cả files cần xử lý"""
        supported_extensions = ['.pdf', '.png', '.jpg', '.jpeg']
        files = []
        
        for file_path in self.data_folder.iterdir():
            if file_path.is_file() and file_path.suffix.lower() in supported_extensions:
                files.append(file_path)
        
        logger.info(f"Tìm thấy {len(files)} files để xử lý")
        return sorted(files)
    
    def extract_from_single_file(self, file_path: Path) -> Dict[str, Any]:
        """Trích xuất thông tin từ một file duy nhất"""
        start_time = time.time()
        
        try:
            # Đọc file và convert nếu cần thiết
            file_data = self._prepare_file_for_processing(file_path)
            
            # Trích xuất sử dụng LangGraphService
            success, extracted_data, error_message = langgraph_service.extract_contract_data_from_file(
                file_data=file_data,
                file_type=file_path.suffix
            )
            
            processing_time = time.time() - start_time
            
            # Tạo kết quả
            result = {
                'file_name': file_path.name,
                'file_path': str(file_path),
                'file_size_mb': file_path.stat().st_size / (1024 * 1024),
                'file_type': file_path.suffix,
                'processing_time_seconds': round(processing_time, 2),
                'success': success,
                'error_message': error_message if not success else None,
                'extraction_timestamp': datetime.now().isoformat(),
            }
            
            # Thêm dữ liệu đã trích xuất nếu thành công
            if success and extracted_data:
                result.update({
                    'start_date': extracted_data.StartDate,
                    'end_date': extracted_data.EndDate,
                    'provider': extracted_data.Provider,
                    'service': extracted_data.Service,
                    'renewal_status': extracted_data.RenewalStatus,
                    'price': getattr(extracted_data, 'Price', None),
                    'currency': getattr(extracted_data, 'Currency', None),
                    'summary_contract': extracted_data.SummaryContract,
                })
                
                # Đánh giá chất lượng dữ liệu
                result['data_quality_score'] = self._calculate_data_quality_score(extracted_data)
            else:
                # Thiết lập giá trị mặc định nếu thất bại
                result.update({
                    'start_date': None,
                    'end_date': None,
                    'provider': None,
                    'service': None,
                    'renewal_status': None,
                    'price': None,
                    'currency': None,
                    'summary_contract': None,
                    'data_quality_score': 0
                })
            
            logger.info(f"Xử lý {'thành công' if success else 'thất bại'} file: {file_path.name} ({processing_time:.2f}s)")
            return result
            
        except Exception as e:
            processing_time = time.time() - start_time
            error_msg = f"Lỗi xử lý file {file_path.name}: {str(e)}"
            logger.error(error_msg)
            
            return {
                'file_name': file_path.name,
                'file_path': str(file_path),
                'file_size_mb': file_path.stat().st_size / (1024 * 1024),
                'file_type': file_path.suffix,
                'processing_time_seconds': round(processing_time, 2),
                'success': False,
                'error_message': error_msg,
                'extraction_timestamp': datetime.now().isoformat(),
                'start_date': None,
                'end_date': None,
                'provider': None,
                'service': None,
                'renewal_status': None,
                'price': None,
                'currency': None,
                'summary_contract': None,
                'data_quality_score': 0
            }
    
    def _prepare_file_for_processing(self, file_path: Path) -> bytes:
        """Chuẩn bị file để xử lý - convert PDF thành image nếu cần"""
        try:
            with open(file_path, 'rb') as f:
                file_data = f.read()
            
            # Nếu là PDF, convert thành image
            if file_path.suffix.lower() == '.pdf':
                try:
                    # Convert PDF thành images
                    images = convert_from_bytes(file_data, first_page=1, last_page=1, dpi=200)
                    
                    if images:
                        # Lấy trang đầu tiên
                        image = images[0]
                        
                        # Convert thành bytes
                        img_buffer = io.BytesIO()
                        image.save(img_buffer, format='PNG')
                        img_buffer.seek(0)
                        
                        return img_buffer.getvalue()
                    else:
                        logger.warning(f"Không thể convert PDF {file_path.name} thành image")
                        return file_data
                        
                except Exception as e:
                    logger.warning(f"Lỗi convert PDF {file_path.name}: {e}, sử dụng raw data")
                    return file_data
            
            # Với image files, trả về như cũ
            return file_data
            
        except Exception as e:
            logger.error(f"Lỗi đọc file {file_path.name}: {e}")
            raise
    
    def _calculate_data_quality_score(self, extracted_data: ExtractedContractData) -> float:
        """Tính điểm chất lượng dữ liệu từ 0-100"""
        score = 0
        
        # Kiểm tra từng trường và tính điểm (tổng 100 điểm)
        if extracted_data.StartDate and extracted_data.StartDate.strip():
            score += 15  # Ngày bắt đầu: 15 điểm
        if extracted_data.EndDate and extracted_data.EndDate.strip():
            score += 15  # Ngày kết thúc: 15 điểm
        if extracted_data.Provider and extracted_data.Provider.strip():
            score += 20  # Nhà cung cấp: 20 điểm
        if extracted_data.Service and extracted_data.Service.strip():
            score += 20  # Dịch vụ: 20 điểm
        
        # Safe access to new fields
        price = getattr(extracted_data, 'Price', None)
        if price and str(price).strip():
            score += 15  # Giá tiền: 15 điểm
        
        currency = getattr(extracted_data, 'Currency', None)  
        if currency and str(currency).strip():
            score += 5   # Tiền tệ: 5 điểm
            
        if extracted_data.RenewalStatus and extracted_data.RenewalStatus.strip():
            score += 5   # Trạng thái gia hạn: 5 điểm
        if extracted_data.SummaryContract and extracted_data.SummaryContract.strip():
            score += 5   # Tóm tắt: 5 điểm
            
        return score
    
    def evaluate_all_files(self) -> None:
        """Đánh giá tất cả files trong thư mục data"""
        files = self.get_file_list()
        self.performance_metrics['total_files'] = len(files)
        
        logger.info(f"Bắt đầu xử lý {len(files)} files...")
        
        for i, file_path in enumerate(files, 1):
            logger.info(f"Đang xử lý file {i}/{len(files)}: {file_path.name}")
            
            result = self.extract_from_single_file(file_path)
            self.results.append(result)
            
            # Cập nhật metrics
            if result['success']:
                self.performance_metrics['successful_extractions'] += 1
            else:
                self.performance_metrics['failed_extractions'] += 1
            
            self.performance_metrics['total_processing_time'] += result['processing_time_seconds']
            
            # In tiến độ mỗi 10 files
            if i % 10 == 0:
                logger.info(f"Đã hoàn thành {i}/{len(files)} files")
        
        # Tính toán metrics cuối cùng
        self._calculate_final_metrics()
        logger.info("Hoàn thành xử lý tất cả files!")
    
    def _calculate_final_metrics(self) -> None:
        """Tính toán các metrics hiệu suất cuối cùng"""
        total_files = self.performance_metrics['total_files']
        successful = self.performance_metrics['successful_extractions']
        
        if total_files > 0:
            self.performance_metrics['success_rate'] = round((successful / total_files) * 100, 2)
            self.performance_metrics['average_processing_time'] = round(
                self.performance_metrics['total_processing_time'] / total_files, 2
            )
        
        # Tính thêm các metrics khác
        if self.results:
            processing_times = [r['processing_time_seconds'] for r in self.results]
            quality_scores = [r['data_quality_score'] for r in self.results if r['success']]
            
            self.performance_metrics.update({
                'min_processing_time': round(min(processing_times), 2),
                'max_processing_time': round(max(processing_times), 2),
                'average_quality_score': round(sum(quality_scores) / len(quality_scores), 2) if quality_scores else 0,
                'files_with_complete_data': len([r for r in self.results if r['data_quality_score'] == 100]),
                'files_with_partial_data': len([r for r in self.results if 0 < r['data_quality_score'] < 100]),
                'files_with_no_data': len([r for r in self.results if r['data_quality_score'] == 0])
            })
    
    def export_to_excel(self, output_file: str = "contract_extraction_evaluation.xlsx") -> None:
        """Xuất kết quả ra file Excel với định dạng đẹp"""
        logger.info(f"Đang xuất kết quả ra file: {output_file}")
        
        # Tạo workbook
        wb = Workbook()
        
        # Sheet 1: Tổng quan hiệu suất
        self._create_performance_summary_sheet(wb)
        
        # Sheet 2: Kết quả chi tiết
        self._create_detailed_results_sheet(wb)
        
        # Sheet 3: Phân tích dữ liệu
        self._create_data_analysis_sheet(wb)
        
        # Lưu file
        wb.save(output_file)
        logger.info(f"Đã xuất kết quả thành công ra file: {output_file}")
    
    def _create_performance_summary_sheet(self, wb: Workbook) -> None:
        """Tạo sheet tổng quan hiệu suất"""
        ws = wb.active
        ws.title = "Tổng quan hiệu suất"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Tiêu đề
        ws["A1"] = "BÁO CÁO ĐÁNH GIÁ HIỆU SUẤT TRÍCH XUẤT THÔNG TIN HỢP ĐỒNG"
        ws["A1"].font = Font(bold=True, size=16)
        
        # Thông tin tổng quan
        ws["A3"] = "Chỉ số"
        ws["B3"] = "Giá trị"
        
        for cell in ["A3", "B3"]:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
            ws[cell].alignment = center_alignment
        
        # Dữ liệu metrics
        metrics_data = [
            ("Tổng số files", self.performance_metrics['total_files']),
            ("Files xử lý thành công", self.performance_metrics['successful_extractions']),
            ("Files xử lý thất bại", self.performance_metrics['failed_extractions']),
            ("Tỷ lệ thành công (%)", self.performance_metrics['success_rate']),
            ("Thời gian xử lý trung bình (giây)", self.performance_metrics['average_processing_time']),
            ("Thời gian xử lý tổng (giây)", self.performance_metrics['total_processing_time']),
            ("Thời gian xử lý min (giây)", self.performance_metrics.get('min_processing_time', 0)),
            ("Thời gian xử lý max (giây)", self.performance_metrics.get('max_processing_time', 0)),
            ("Điểm chất lượng dữ liệu trung bình", self.performance_metrics.get('average_quality_score', 0)),
            ("Files có dữ liệu đầy đủ", self.performance_metrics.get('files_with_complete_data', 0)),
            ("Files có dữ liệu một phần", self.performance_metrics.get('files_with_partial_data', 0)),
            ("Files không có dữ liệu", self.performance_metrics.get('files_with_no_data', 0)),
        ]
        
        for i, (metric, value) in enumerate(metrics_data, 4):
            ws[f"A{i}"] = metric
            ws[f"B{i}"] = value
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
    
    def _create_detailed_results_sheet(self, wb: Workbook) -> None:
        """Tạo sheet kết quả chi tiết"""
        ws = wb.create_sheet("Kết quả chi tiết")
        
        # Chuyển results thành DataFrame
        df = pd.DataFrame(self.results)
        
        # Thêm dữ liệu vào sheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Styling header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-fit columns - cách an toàn hơn
        for i, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = ws.cell(row=1, column=i).column_letter
            
            for cell in column:
                try:
                    if hasattr(cell, 'value') and cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_data_analysis_sheet(self, wb: Workbook) -> None:
        """Tạo sheet phân tích dữ liệu"""
        ws = wb.create_sheet("Phân tích dữ liệu")
        
        # Tiêu đề
        ws["A1"] = "PHÂN TÍCH CHI TIẾT DỮ LIỆU TRÍCH XUẤT"
        ws["A1"].font = Font(bold=True, size=14)
        
        # Phân tích theo loại file
        file_types = {}
        for result in self.results:
            file_type = result['file_type']
            if file_type not in file_types:
                file_types[file_type] = {'total': 0, 'success': 0}
            
            file_types[file_type]['total'] += 1
            if result['success']:
                file_types[file_type]['success'] += 1
        
        ws["A3"] = "Phân tích theo loại file:"
        ws["A3"].font = Font(bold=True)
        
        ws["A4"] = "Loại file"
        ws["B4"] = "Tổng số"
        ws["C4"] = "Thành công"
        ws["D4"] = "Tỷ lệ thành công (%)"
        
        for cell in ["A4", "B4", "C4", "D4"]:
            ws[cell].font = Font(bold=True)
        
        row = 5
        for file_type, stats in file_types.items():
            success_rate = (stats['success'] / stats['total']) * 100 if stats['total'] > 0 else 0
            ws[f"A{row}"] = file_type
            ws[f"B{row}"] = stats['total']
            ws[f"C{row}"] = stats['success']
            ws[f"D{row}"] = round(success_rate, 2)
            row += 1
        
        # Phân tích theo chất lượng dữ liệu
        ws[f"A{row+2}"] = "Phân tích chất lượng dữ liệu:"
        ws[f"A{row+2}"].font = Font(bold=True)
        
        quality_ranges = {"0-20": 0, "21-40": 0, "41-60": 0, "61-80": 0, "81-100": 0}
        
        for result in self.results:
            score = result['data_quality_score']
            if score <= 20:
                quality_ranges["0-20"] += 1
            elif score <= 40:
                quality_ranges["21-40"] += 1
            elif score <= 60:
                quality_ranges["41-60"] += 1
            elif score <= 80:
                quality_ranges["61-80"] += 1
            else:
                quality_ranges["81-100"] += 1
        
        ws[f"A{row+3}"] = "Khoảng điểm"
        ws[f"B{row+3}"] = "Số lượng files"
        
        for cell in [f"A{row+3}", f"B{row+3}"]:
            ws[cell].font = Font(bold=True)
        
        row += 4
        for range_name, count in quality_ranges.items():
            ws[f"A{row}"] = range_name
            ws[f"B{row}"] = count
            row += 1
        
        # Auto-fit columns - cách an toàn hơn  
        for i, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = ws.cell(row=1, column=i).column_letter
            
            for cell in column:
                try:
                    if hasattr(cell, 'value') and cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def print_summary(self) -> None:
        """In tóm tắt kết quả đánh giá"""
        print("\n" + "="*80)
        print("           BÁO CÁO ĐÁNH GIÁ HIỆU SUẤT TRÍCH XUẤT THÔNG TIN")
        print("="*80)
        
        for key, value in self.performance_metrics.items():
            print(f"{key.replace('_', ' ').title():<30}: {value}")
        
        print("="*80 + "\n")


def main():
    """Hàm chính chạy evaluation"""
    try:
        # Khởi tạo evaluator
        evaluator = ContractExtractionEvaluator("data")
        
        # Chạy đánh giá
        evaluator.evaluate_all_files()
        
        # Xuất kết quả
        evaluator.export_to_excel("contract_extraction_evaluation_results.xlsx")
        
        # In tóm tắt
        evaluator.print_summary()
        
        # Lưu kết quả dưới dạng JSON để backup
        with open("evaluation_results.json", "w", encoding="utf-8") as f:
            json.dump({
                'performance_metrics': evaluator.performance_metrics,
                'detailed_results': evaluator.results
            }, f, ensure_ascii=False, indent=2)
        
        logger.info("Đánh giá hoàn thành! Các file output:")
        logger.info("- contract_extraction_evaluation_results.xlsx")
        logger.info("- evaluation_results.json")
        logger.info("- evaluation_log.txt")
        
    except Exception as e:
        logger.error(f"Lỗi khi chạy evaluation: {e}")
        raise


if __name__ == "__main__":
    main()
